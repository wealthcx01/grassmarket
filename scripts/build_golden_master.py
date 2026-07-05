"""Build the ATLAS golden-master fixture (GRS-0003).

Computes ONE complete assessment end-to-end per ATLAS Methodology v1 §5, for a composite
mid-tier retail brokerage ("Meridian Securities"), and emits:

  - tests/fixtures/golden_master.json  — inputs + every intermediate + final value (authoritative
    oracle the engine must reproduce exactly in GRS-0004).
  - fixtures/golden-master.xlsx        — the same case with VISIBLE, LIVE spreadsheet formulas so
    John can audit and adjust the judgement calls. Editing an input level recomputes downstream.

This generator is a REFERENCE implementation of the methodology, not the engine. The engine
(GRS-0004) is written independently and must reproduce this JSON to the last decimal; the property
tests provide the independent cross-check. After John edits the workbook in Excel and saves it,
`scripts/regen_golden_master_json.py` reads the recalculated values back into the JSON.

Fail-loud: the inputs must cover EXACTLY the registry's 51 subcomponents, 10 metrics, and 7
powers — a missing or extra key aborts the build (the same discipline ADR-0001 enforces at run
time). All coefficients here are DRAFT (uniform, documented) pending the elicitation panel; the
strength encoding and the rating-gate rule are draft interpretations flagged for ratification.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from bcap_contracts import load_registry
from bcap_contracts.common import NonScoreState
from bcap_contracts.registry import MetricDef, Registry
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

_ROOT = Path(__file__).resolve().parents[1]
_JSON_OUT = _ROOT / "tests" / "fixtures" / "golden_master.json"
_XLSX_OUT = _ROOT / "fixtures" / "golden-master.xlsx"

METHODOLOGY_VERSION = "1.1"  # folds the golden-master extensions into the Methodology (ADR-0003..7)
FIXTURE_STATUS = "draft-pending-ratification"
SUBJECT = "Meridian Securities (composite mid-tier retail brokerage)"

# --- Scale + encodings ------------------------------------------------------------------
LEVEL_INDEX = {"Basic": 0.2, "Developing": 0.5, "Advanced": 0.8, "Frontier": 1.0}
LEVEL_RANK = {"Basic": 1, "Developing": 2, "Advanced": 3, "Frontier": 4}
EVIDENCE_RANK = {"E1": 1, "E2": 2, "E3": 3, "E4": 4}
# The non-score state strings are the CONTRACT enum VALUES (D8 single vocabulary) — the same
# strings the engine emits, so the fixture is a contract-vocabulary oracle (GRS-0005 precursor).
NOT_APPLICABLE = NonScoreState.NOT_APPLICABLE.value  # "Not Applicable"
NOT_ASSESSED = NonScoreState.NOT_ASSESSED.value  # "Not Assessed"

# Numeric encoding of the ordinal power strength for the continuous P index (ADR-0004). The triad
# ratings themselves stay ordinal (ADR-0002); this feeds only P.
STRENGTH_VALUE = {"None": 0.0, "Emerging": 0.4, "Established": 0.7, "Wide": 1.0}
STRENGTH_RANK = {"None": 0, "Emerging": 1, "Established": 2, "Wide": 3}
# numeric-aggregate → ordinal band thresholds for the derived triad ratings (ADR-0007, accepted).
# These are NOT free parameters: each is the MIDPOINT between adjacent STRENGTH_VALUE anchors, i.e.
# nearest-named-level discretisation — (0.0+0.4)/2=0.2, (0.4+0.7)/2=0.55, (0.7+1.0)/2=0.85. If the
# encoding is re-elicited, regenerate these midpoints from it rather than tuning them independently.
_TRIAD_THRESHOLDS = (("Wide", 0.85), ("Established", 0.55), ("Emerging", 0.2), ("None", 0.0))

# Draft group weights for the B index (ADR-0006): scale metrics count as ONE group, not ~4× (B1).
GROUP_WEIGHTS = {"scale": 1.0, "unit_economics": 1.0, "momentum": 1.0}

# --- DRAFT coefficients (uniform; pending elicitation) ----------------------------------
ALPHA_MODULE = 0.7  # per-module blend α (uniform draft)
ALPHA_L = 0.7  # L-level blend α_L (draft)
THETA = {"B": 0.30, "P": 0.30, "L": 0.40}  # Σθ = 1 (draft; L weighted as the assessment core)
CRITICAL_MODULES_FOR_L = ("APP_SERVER", "BACKOFFICE", "OEMS")  # draft: brokerage core
# λ (subcomponent), δ (module), w_metric, w_power are all uniform 1.0 in this draft fixture.

# --- Subject inputs ---------------------------------------------------------------------
# Each subcomponent: (level, evidence) if assessed, or (state, None). Criticals are all assessed
# here for a clean golden number; non-score states sit in non-critical slots. Two Not-Assessed
# and one Not-Applicable case exercise §3.2.
SUBCOMPONENTS: dict[str, tuple[str, str | None]] = {
    # FRONTEND
    "FRONTEND_PERFORMANCE": ("Advanced", "E3"),
    "FRONTEND_UX_NAVIGATION": ("Advanced", "E4"),
    "FRONTEND_DEVICE_COVERAGE": ("Developing", "E2"),
    "FRONTEND_PERSONALISATION": ("Developing", "E2"),
    "FRONTEND_ACCESSIBILITY_LOCALISATION": ("Basic", "E2"),
    "FRONTEND_EXPERIMENTATION_ANALYTICS": ("Developing", "E3"),
    # APP_SERVER
    "APP_SERVER_HOSTING_ELASTICITY": ("Advanced", "E3"),
    "APP_SERVER_RESILIENCE_DR": ("Developing", "E3"),
    "APP_SERVER_API_DESIGN": ("Advanced", "E3"),
    "APP_SERVER_SECURITY_COMPLIANCE": ("Advanced", "E4"),
    "APP_SERVER_DEVOPS_DEPLOYMENT": ("Developing", "E2"),
    "APP_SERVER_DATA_ARCHITECTURE": ("Developing", "E2"),
    "APP_SERVER_OBSERVABILITY": ("Advanced", "E3"),
    # MARKET_DATA
    "MARKET_DATA_EXCHANGE_COVERAGE": ("Advanced", "E3"),
    "MARKET_DATA_INSTRUMENT_UNIVERSE": ("Advanced", "E2"),
    "MARKET_DATA_DEPTH_QUALITY": ("Advanced", "E3"),
    "MARKET_DATA_LATENCY_TIMELINESS": ("Developing", "E3"),
    "MARKET_DATA_HISTORY_DEPTH": ("Developing", "E2"),
    "MARKET_DATA_VENDOR_REDUNDANCY": ("Basic", "E2"),
    "MARKET_DATA_VALUE_ADD_SERVICES": (NOT_ASSESSED, None),
    # ORCHESTRATION
    "ORCHESTRATION_WORKFLOW_ENGINE": ("Developing", "E3"),
    "ORCHESTRATION_ROUTING_LOGIC": ("Developing", "E2"),
    "ORCHESTRATION_EVENT_DRIVEN": ("Basic", "E2"),
    "ORCHESTRATION_CONFIG_VS_CODE": ("Developing", "E2"),
    "ORCHESTRATION_MONITORING": ("Developing", "E3"),
    # CMS
    "CMS_RESEARCH_AUTHORING": ("Advanced", "E3"),
    "CMS_RESEARCH_DISTRIBUTION": ("Developing", "E2"),
    "CMS_EMAIL_CAMPAIGNS": ("Advanced", "E3"),
    "CMS_CRM_MODEL": ("Developing", "E3"),
    "CMS_STATEMENTS": ("Advanced", "E3"),
    "CMS_CONTENT_SEARCH_PERSONALISATION": ("Basic", "E2"),
    # BACKOFFICE
    "BACKOFFICE_CUSTODY": ("Advanced", "E4"),
    "BACKOFFICE_PAYMENTS_FUNDING": ("Advanced", "E3"),
    "BACKOFFICE_KYC_ONBOARDING": ("Advanced", "E3"),
    "BACKOFFICE_PORTFOLIO_MGMT": ("Developing", "E2"),
    "BACKOFFICE_CREDIT_RISK": ("Developing", "E2"),
    "BACKOFFICE_REG_REPORTING": ("Advanced", "E4"),
    # OEMS
    "OEMS_ASSET_COVERAGE": ("Developing", "E2"),
    "OEMS_EXEC_ALGOS": ("Basic", "E2"),
    "OEMS_PRE_TRADE_RISK": ("Advanced", "E3"),
    "OEMS_ORDER_TYPES": ("Developing", "E2"),
    "OEMS_LATENCY_RELIABILITY": ("Developing", "E3"),
    "OEMS_APIS_COLOCATION": (NOT_APPLICABLE, None),
    # EMS_GATEWAY
    "EMS_GATEWAY_CONNECTIVITY": ("Developing", "E2"),
    "EMS_GATEWAY_ROUTING_POLICY": ("Developing", "E2"),
    "EMS_GATEWAY_RISK_THROTTLING": ("Advanced", "E3"),
    "EMS_GATEWAY_MONITORING": ("Developing", "E2"),
    # LIQ_CONNECT
    "LIQ_CONNECT_LOCAL_EXCHANGES": ("Advanced", "E3"),
    "LIQ_CONNECT_FOREIGN_BROKERS": ("Developing", "E2"),
    "LIQ_CONNECT_FUND_HOUSES": (NOT_ASSESSED, None),
    "LIQ_CONNECT_SETTLEMENT_CLEARING": ("Advanced", "E3"),
}

# Business metric raw values (declared unit), OR a non-score state string (NOT_APPLICABLE /
# NOT_ASSESSED) — metrics now support §3.2 states with renormalisation (review B4).
METRIC_RAW: dict[str, float | str] = {
    "AUA": 1_800_000_000,
    "ACTIVE_CLIENTS": 180_000,
    "NET_REVENUE": 85_000_000,
    "REVENUE_PER_CLIENT": 472,
    "GROSS_MARGIN": 58,
    "COST_TO_SERVE": 140,
    "NET_REVENUE_RETENTION": 104,
    "CLIENT_GROWTH_RATE": 14,
    "TAKE_RATE_LEVEL": 22,  # renamed from TAKE_RATE_DURABILITY (review B3)
    "CAC_PAYBACK_MONTHS": 18,
}

# Powers carry a BENEFIT strength and a BARRIER strength (Helmer: a power needs both; review B8).
# Per-power strength = the weaker of Benefit/Barrier; the triad (§2) is derived from the two sides.
# N/A is NOT a legal state for a power — all 7 powers are always in scope (Methodology v1.1 §8;
# ADR-0007 revised). A power that is structurally weak for this business scores a real LOW level,
# not N/A: absent-because-weak is a genuine None/Emerging reading, and Helmer's both-required test
# already collapses a one-sided power to its weaker (often None) side.
#
# Meridian (draft, John to ratify): Network Economies is weak — some social/referral sharing gives
# an Emerging benefit, but there is no installed-base lock-in, so the Barrier is None (a challenger
# need not match a network to compete). Counter-Positioning is weak vs traditional banks — the
# digital, low-fee model is a genuine Emerging benefit, but incumbents are not structurally deterred
# from launching rival digital brokerages, so the Barrier is None. Both therefore score strength
# None (the weaker side) — real, low, and honest rather than dropped.
POWER_ASSESSMENT: dict[str, tuple[str, str]] = {
    "SCALE_ECONOMIES": ("Emerging", "Emerging"),
    "NETWORK_ECONOMIES": ("Emerging", "None"),
    "COUNTER_POSITIONING": ("Emerging", "None"),
    "SWITCHING_COSTS": ("Established", "Established"),
    "BRANDING": ("Emerging", "Emerging"),
    "CORNERED_RESOURCE": ("None", "None"),
    "PROCESS_POWER": ("Emerging", "Emerging"),
}


def _round(x: float) -> float:
    return round(x, 6)


def _interpolate(metric: MetricDef, raw: float) -> float:
    """Piecewise-linear normalisation n_k(raw) → [0,1] against the metric's anchors (§5.3).
    Anchors are ordered by ascending raw; values are clamped outside the anchor range."""
    anchors = metric.normalisation.anchors
    if not anchors:
        raise ValueError(f"Metric {metric.key} has no normalisation anchors.")
    pts = sorted(((a.raw, a.normalised) for a in anchors), key=lambda p: p[0])
    if raw <= pts[0][0]:
        return pts[0][1]
    if raw >= pts[-1][0]:
        return pts[-1][1]
    for (x0, y0), (x1, y1) in zip(pts, pts[1:], strict=False):
        if x0 <= raw <= x1:
            t = (raw - x0) / (x1 - x0)
            return y0 + t * (y1 - y0)
    raise AssertionError("unreachable")  # pragma: no cover


_RANK_BAND = {1: "Basic", 2: "Developing", 3: "Advanced", 4: "Frontier"}


def _evidence_rank(evidence: str | None) -> int:
    """Fail-loud evidence lookup — an assessed subcomponent that reaches the gate MUST carry an
    evidence grade. No `ev or "E1"` default (that was the banned defensive-default species, A6)."""
    if evidence is None:
        raise ValueError("Assessed subcomponent reached the gate without an evidence grade.")
    return EVIDENCE_RANK[evidence]


def _rating_gate(
    critical_inputs: list[tuple[str, str | None]], all_assessed_ranks: list[int]
) -> tuple[str, bool, str]:
    """Rule-based rating gate (§5.2). DRAFT interpretation, flagged for ratification.

    band = min(critical-rule CEILING, overall-maturity FLOOR):
      - CEILING (necessary conditions on CRITICAL subcomponents):
          * a critical Not Assessed  → blocked, ceiling = Developing (cannot certify Advanced+);
          * all critical Advanced+ at E3+ → Frontier;
          * no critical is Basic      → Advanced;
          * otherwise (a critical is Basic) → Developing.
      - FLOOR (bottleneck over ALL assessed subcomponents): the module cannot be rated above its
        weakest assessed part — any Basic caps at Developing, min Developing caps at Advanced.
    N/A subcomponents are out of scope (dropped). This keeps the headline honest: a module with a
    Basic part is never Frontier, even if its critical subcomponents are strong.
    """
    assessed_crit = [(lvl, ev) for (lvl, ev) in critical_inputs if lvl in LEVEL_INDEX]
    blocked = any(lvl == NOT_ASSESSED for (lvl, _) in critical_inputs)
    crit_ranks = [LEVEL_RANK[lvl] for (lvl, _) in assessed_crit]
    if not assessed_crit and not blocked:
        ceiling_rank, ceiling_note = 4, "no critical subcomponent in scope"
    elif blocked:
        ceiling_rank, ceiling_note = 2, "gate blocked: a critical subcomponent is Not Assessed"
    elif all(LEVEL_RANK[lvl] >= 3 and _evidence_rank(ev) >= 3 for (lvl, ev) in assessed_crit):
        ceiling_rank, ceiling_note = 4, "all critical Advanced+ at E3+"
    elif min(crit_ranks) >= 2:
        ceiling_rank, ceiling_note = 3, "no critical is Basic"
    elif max(crit_ranks) == 1:
        ceiling_rank, ceiling_note = 1, "every critical is Basic"
    else:
        ceiling_rank, ceiling_note = 2, "a critical subcomponent is Basic"

    # FLOOR (overall bottleneck): all assessed Advanced+ permits Frontier; a Developing minimum
    # caps at Advanced; a single Basic caps at Developing; ALL-Basic floors the band at Basic.
    floor_rank = min(all_assessed_ranks) if all_assessed_ranks else 1
    if floor_rank >= 3:
        floor_cap = 4
    elif floor_rank == 2:
        floor_cap = 3
    elif all(r == 1 for r in all_assessed_ranks):
        floor_cap = 1  # every assessed subcomponent is Basic → the module is Basic (now reachable)
    else:
        floor_cap = 2

    band_rank = min(ceiling_rank, floor_cap)
    if band_rank == ceiling_rank and ceiling_rank <= floor_cap:
        note = ceiling_note
    else:
        note = f"{ceiling_note}; capped by bottleneck ({_RANK_BAND[floor_rank]})"
    return (_RANK_BAND[band_rank], blocked, note)


def _to_ordinal(value: float) -> str:
    """Map a numeric aggregate in [0,1] to an ordinal triad rating (ADR-0007, draft thresholds)."""
    for label, threshold in _TRIAD_THRESHOLDS:
        if value >= threshold:
            return label
    return "None"  # pragma: no cover — thresholds include 0.0


def _compute_triad(
    power_rows: list[dict[str, Any]], group_means: dict[str, float]
) -> dict[str, Any]:
    """Derive the Platform Power triad (Methodology §2) from the split Benefit/Barrier power data.

    Ordinal OUT (ADR-0002): each dimension reports a `rating`; the numeric `score` is shown only for
    audit. Economic = B's scale + unit-economics signal; Perceived = Benefit-side of Branding +
    Switching Costs; Defence = the Barrier-side aggregate across ALL 7 powers (the moat) — powers
    are never N/A (Methodology v1.1 §8), so this is the literal §2 across-all-powers aggregate."""
    barriers = [STRENGTH_VALUE[p["barrier"]] for p in power_rows]
    defence = sum(barriers) / len(barriers)
    perceived_src = [p for p in power_rows if p["key"] in ("BRANDING", "SWITCHING_COSTS")]
    perceived = (
        sum(STRENGTH_VALUE[p["benefit"]] for p in perceived_src) / len(perceived_src)
        if perceived_src
        else 0.0
    )
    econ_src = [group_means[g] for g in ("scale", "unit_economics") if g in group_means]
    economic = sum(econ_src) / len(econ_src) if econ_src else 0.0
    return {
        "economic_value": {"rating": _to_ordinal(economic), "score": _round(economic)},
        "perceived_value": {"rating": _to_ordinal(perceived), "score": _round(perceived)},
        "defence_value": {"rating": _to_ordinal(defence), "score": _round(defence)},
    }


def compute(registry: Registry) -> dict[str, Any]:
    """Compute the whole assessment. Fail-loud on any input/registry mismatch."""
    _assert_inputs_cover_registry(registry)

    modules_out: list[dict[str, Any]] = []
    q_by_module: dict[str, float] = {}
    for module in registry.modules:
        sub_rows: list[dict[str, Any]] = []
        applicable_assessed: list[tuple[str, str, str | None]] = []  # (key, level, evidence)
        critical_inputs: list[tuple[str, str | None]] = []
        n_applicable = n_assessed = n_na = 0
        for sub in module.subcomponents:
            level, evidence = SUBCOMPONENTS[sub.key]
            state = level if level in (NOT_APPLICABLE, NOT_ASSESSED) else None
            row = {
                "key": sub.key,
                "critical": sub.critical,
                "level": None if state else level,
                "index": None if state else LEVEL_INDEX[level],
                "evidence": evidence,
                "state": state,
            }
            sub_rows.append(row)
            if state == NOT_APPLICABLE:
                n_na += 1
            else:
                n_applicable += 1
                if state == NOT_ASSESSED:
                    pass  # applicable but not assessed — excluded from q_m, taints gate/coverage
                else:
                    n_assessed += 1
                    applicable_assessed.append((sub.key, level, evidence))
            if sub.critical:
                critical_inputs.append((level, evidence))

        # q_m over APPLICABLE + ASSESSED only (λ uniform = 1.0 → weighted avg is simple mean).
        indices = [LEVEL_INDEX[lvl] for (_, lvl, _) in applicable_assessed]
        if indices:
            weighted_avg = sum(indices) / len(indices)
            min_term = min(indices)
            bottleneck = min(applicable_assessed, key=lambda t: LEVEL_INDEX[t[1]])[0]
            q_m = ALPHA_MODULE * weighted_avg + (1 - ALPHA_MODULE) * min_term
        else:
            weighted_avg = min_term = q_m = None
            bottleneck = None
        all_assessed_ranks = [LEVEL_RANK[lvl] for (_, lvl, _) in applicable_assessed]
        band, blocked, note = _rating_gate(critical_inputs, all_assessed_ranks)
        coverage = n_assessed / n_applicable if n_applicable else None

        q_by_module[module.key] = q_m  # None if the module is fully unassessed — never 0.0 (D9)
        modules_out.append(
            {
                "key": module.key,
                "name": module.name,
                "subcomponents": sub_rows,
                "n_applicable": n_applicable,
                "n_assessed": n_assessed,
                "n_not_applicable": n_na,
                "coverage": _round(coverage) if coverage is not None else None,
                "alpha": ALPHA_MODULE,
                "weighted_avg": _round(weighted_avg) if weighted_avg is not None else None,
                "min_term": _round(min_term) if min_term is not None else None,
                "bottleneck_subcomponent": bottleneck,
                "q_m": _round(q_m) if q_m is not None else None,
                "gate_band": band,
                "gate_blocked": blocked,
                "gate_note": note,
            }
        )

    # L = α_L·(Σδ·q_m/Σδ) + (1−α_L)·min(q_critical_modules). δ uniform = 1.0. A fully-unassessed
    # module (q_m None) is EXCLUDED from both terms — never zero-filled (D9). δ renormalises over
    # the assessed modules; the critical-module min ranges only over assessed critical modules.
    assessed_q = {k: v for k, v in q_by_module.items() if v is not None}
    if not assessed_q:
        raise ValueError("Cannot compute L: no module has any assessed subcomponent.")
    l_weighted = sum(assessed_q.values()) / len(assessed_q)
    crit_q = [q_by_module[k] for k in CRITICAL_MODULES_FOR_L if q_by_module[k] is not None]
    if not crit_q:
        raise ValueError("Cannot compute L min term: no critical-for-L module is assessed.")
    l_min = min(crit_q)
    L = ALPHA_L * l_weighted + (1 - ALPHA_L) * l_min

    # B = group-weighted mean (ADR-0006). Each metric's n_k feeds its group's mean; groups are
    # then combined with GROUP_WEIGHTS so collinear scale metrics count once (review B1). Metrics
    # in a NOT_APPLICABLE / NOT_ASSESSED state are excluded and their group renormalises (B4).
    metric_rows: list[dict[str, Any]] = []
    group_ns: dict[str, list[float]] = {}
    for metric in registry.metrics:
        raw = METRIC_RAW[metric.key]
        state = raw if isinstance(raw, str) else None
        n_k = None if state else _interpolate(metric, float(raw))
        metric_rows.append(
            {
                "key": metric.key,
                "raw": None if state else raw,
                "unit": metric.unit,
                "direction": metric.direction,
                "group": metric.group,
                "state": state,
                "n_k": _round(n_k) if n_k is not None else None,
            }
        )
        if n_k is not None:
            group_ns.setdefault(metric.group or "ungrouped", []).append(n_k)
    group_means = {g: sum(v) / len(v) for g, v in group_ns.items()}
    b_num = sum(GROUP_WEIGHTS[g] * mean for g, mean in group_means.items())
    b_den = sum(GROUP_WEIGHTS[g] for g in group_means)
    B = b_num / b_den

    # P = mean over ALL 7 powers of each power's strength, where a power's strength is the WEAKER
    # of its Benefit and Barrier (Helmer: a power needs both; review B8). N/A is not a legal power
    # state (Methodology v1.1 §8) — every power is in scope, and a structurally weak power scores a
    # real low strength (often None via the weaker side) rather than being dropped.
    power_rows: list[dict[str, Any]] = []
    power_values: list[float] = []
    for power in registry.powers:
        benefit, barrier = POWER_ASSESSMENT[power.key]
        strength = benefit if STRENGTH_RANK[benefit] <= STRENGTH_RANK[barrier] else barrier
        value = STRENGTH_VALUE[strength]
        power_values.append(value)
        power_rows.append(
            {
                "key": power.key,
                "benefit": benefit,
                "barrier": barrier,
                "strength": strength,  # the weaker side (Helmer both-required)
                "value": value,
                "state": None,  # retained for output-shape stability; always None (no N/A powers)
            }
        )
    P = sum(power_values) / len(power_values)

    triad = _compute_triad(power_rows, group_means)

    V = THETA["B"] * B + THETA["P"] * P + THETA["L"] * L
    # Display is derived from the STORED (rounded) V so 0–100 == V×100 exactly (ADR-0001 §4). Build
    # the convention string from the computed value so it can never go stale (review defect: it once
    # cited a V from a superseded power model).
    v_stored = _round(V)
    v_display = _round(v_stored * 100)

    return {
        "metadata": {
            "fixture": "atlas-golden-master",
            "subject": SUBJECT,
            "methodology_version": METHODOLOGY_VERSION,
            "status": FIXTURE_STATUS,
            "display_convention": (
                f"Indices are stored on [0,1] to 6 dp; the 0–100 display is the STORED V × 100 "
                f"(ADR-0001 §4). Narrative may round further (e.g. {v_display:.2f}) — the fixture "
                f"value ({v_display}) is authoritative."
            ),
            "note": (
                "Reference computation of Methodology v1.1 §5. Coefficients and the power-strength "
                "encoding are DRAFT pending John's ratification and the elicitation panel; the "
                "gate rule, metric grouping, power-strength = min(Benefit, Barrier), and "
                "powers-never-N/A are now Methodology v1.1 (ADR-0003..0007). The engine (GRS-0004) "
                "must reproduce these values exactly."
            ),
        },
        "coefficients": {
            "alpha_module": ALPHA_MODULE,
            "alpha_l": ALPHA_L,
            "theta": THETA,
            "critical_modules_for_l": list(CRITICAL_MODULES_FOR_L),
            "lambda": "uniform 1.0 per subcomponent (draft)",
            "delta": "uniform 1.0 per module (draft)",
            "w_metric": "uniform within group (draft)",
            "group_weights": GROUP_WEIGHTS,
            "w_power": "uniform 1.0 over applicable powers (draft)",
            "strength_encoding": STRENGTH_VALUE,
            "triad_thresholds": {label: t for label, t in _TRIAD_THRESHOLDS},
        },
        "modules": modules_out,
        "L": {"weighted_term": _round(l_weighted), "min_term": _round(l_min), "value": _round(L)},
        "business": {
            "metrics": metric_rows,
            "group_means": {g: _round(m) for g, m in group_means.items()},
            "B": _round(B),
        },
        "powers": {"powers": power_rows, "P": _round(P)},
        "triad": triad,
        "composite": {"B": _round(B), "P": _round(P), "L": _round(L), "V": _round(V)},
        "two_track": {
            # Display is derived from the STORED (rounded) V so the 0–100 figure equals V×100
            # exactly (the display layer only ever scales the stored score — ADR-0001 §4).
            "continuous": {"V": _round(V), "V_display_0_100": _round(_round(V) * 100)},
            "gates": {m["key"]: m["gate_band"] for m in modules_out},
        },
    }


def _assert_inputs_cover_registry(registry: Registry) -> None:
    reg_subs = registry.all_subcomponent_keys()
    given_subs = set(SUBCOMPONENTS)
    if reg_subs != given_subs:
        raise ValueError(
            f"Subcomponent inputs must cover the registry exactly. "
            f"Missing: {sorted(reg_subs - given_subs)}; extra: {sorted(given_subs - reg_subs)}"
        )
    if registry.metric_keys() != set(METRIC_RAW):
        raise ValueError(
            f"Metric inputs must cover the registry exactly. "
            f"Missing: {sorted(registry.metric_keys() - set(METRIC_RAW))}; "
            f"extra: {sorted(set(METRIC_RAW) - registry.metric_keys())}"
        )
    if registry.power_keys() != set(POWER_ASSESSMENT):
        raise ValueError("Power inputs must cover the registry exactly.")


def write_json(result: dict[str, Any]) -> None:
    _JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    _JSON_OUT.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


_HEADER_FILL = PatternFill("solid", fgColor="1A3B26")  # Bottle Green
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1A3B26")
_NOTE_FONT = Font(italic=True, color="666666")


def _header(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def write_workbook(registry: Registry, result: dict[str, Any], path: Path) -> None:
    """Write the audit workbook with LIVE formulas for the numeric pipeline (subcomponent levels
    → q_m → L, metrics → B, powers → P → V). Rating-gate bands are rule-based, non-arithmetic, so
    they are written as computed values with a note. Editing a Level recomputes V."""
    wb = Workbook()

    # --- Subcomponents (source of truth for the levels John ratifies) ---
    ws_sub = wb.active
    ws_sub.title = "Subcomponents"
    _header(
        ws_sub, 1, ["Module", "Subcomponent", "Critical", "Level", "Index", "Evidence", "State"]
    )
    module_ranges: dict[str, tuple[int, int]] = {}
    r = 2
    for module in registry.modules:
        first = r
        for sub in module.subcomponents:
            row = next(m for m in result["modules"] if m["key"] == module.key)
            sub_row = next(s for s in row["subcomponents"] if s["key"] == sub.key)
            ws_sub.cell(row=r, column=1, value=module.key)
            ws_sub.cell(row=r, column=2, value=sub.key)
            ws_sub.cell(row=r, column=3, value="Yes" if sub.critical else "")
            ws_sub.cell(row=r, column=4, value=sub_row["level"] or "")
            ws_sub.cell(
                row=r,
                column=5,
                value=(
                    f'=IF($D{r}="","",IF($D{r}="Basic",0.2,IF($D{r}="Developing",0.5,'
                    f'IF($D{r}="Advanced",0.8,IF($D{r}="Frontier",1,"")))))'
                ),
            )
            ws_sub.cell(row=r, column=6, value=sub_row["evidence"] or "")
            ws_sub.cell(row=r, column=7, value=sub_row["state"] or "")
            r += 1
        module_ranges[module.key] = (first, r - 1)
    _autofit(ws_sub, {1: 16, 2: 30, 4: 12, 7: 16})

    # --- Modules (q_m live from the Subcomponents index ranges) ---
    ws_mod = wb.create_sheet("Modules")
    _header(
        ws_mod,
        1,
        [
            "Module",
            "alpha",
            "weighted_avg",
            "min_term",
            "q_m",
            "gate_band",
            "gate_blocked",
            "coverage",
        ],
    )
    module_qm_cell: dict[str, str] = {}
    mr = 2
    for module in registry.modules:
        first, last = module_ranges[module.key]
        rng = f"Subcomponents!$E${first}:$E${last}"
        state_rng = f"Subcomponents!$G${first}:$G${last}"
        m = next(x for x in result["modules"] if x["key"] == module.key)
        ws_mod.cell(row=mr, column=1, value=module.key)
        ws_mod.cell(row=mr, column=2, value=ALPHA_MODULE)
        ws_mod.cell(row=mr, column=3, value=f'=IFERROR(AVERAGE({rng}),"")')
        ws_mod.cell(row=mr, column=4, value=f"=MIN({rng})")
        ws_mod.cell(row=mr, column=5, value=f"=$B{mr}*$C{mr}+(1-$B{mr})*$D{mr}")
        ws_mod.cell(row=mr, column=6, value=m["gate_band"])
        ws_mod.cell(row=mr, column=7, value="Yes" if m["gate_blocked"] else "")
        ws_mod.cell(
            row=mr,
            column=8,
            value=f'=IFERROR(COUNT({rng})/(COUNT({rng})+COUNTIF({state_rng},"NOT_ASSESSED")),"")',
        )
        module_qm_cell[module.key] = f"Modules!$E${mr}"
        mr += 1
    _autofit(ws_mod, {1: 16, 6: 12})

    # --- Business (B = group-weighted mean; n_k live piecewise; metric states supported) ---
    ws_b = wb.create_sheet("Business")
    _header(
        ws_b,
        1,
        [
            "Metric",
            "Group",
            "Raw",
            "Unit",
            "Direction",
            "x1",
            "x2",
            "x3",
            "x4",
            "y1",
            "y2",
            "y3",
            "y4",
            "n_k",
        ],
    )
    br = 2
    for metric in registry.metrics:
        anchors = sorted(metric.normalisation.anchors, key=lambda a: a.raw)
        if len(anchors) != 4:
            raise ValueError(
                f"Workbook piecewise formula expects 4 anchors; {metric.key} has {len(anchors)}."
            )
        ws_b.cell(row=br, column=1, value=metric.key)
        ws_b.cell(row=br, column=2, value=metric.group)
        ws_b.cell(row=br, column=3, value=METRIC_RAW[metric.key])  # number or state string
        ws_b.cell(row=br, column=4, value=metric.unit)
        ws_b.cell(row=br, column=5, value=metric.direction)
        for i, a in enumerate(anchors):
            ws_b.cell(row=br, column=6 + i, value=a.raw)  # x1..x4 → F..I
            ws_b.cell(row=br, column=10 + i, value=a.normalised)  # y1..y4 → J..M
        # n_k (col N): blank when Raw is a non-numeric state; else piecewise-linear, clamped.
        ws_b.cell(
            row=br,
            column=14,
            value=(
                f'=IF(NOT(ISNUMBER($C{br})),"",'
                f"IF($C{br}<=$F{br},$J{br},IF($C{br}>=$I{br},$M{br},"
                f"IF($C{br}<=$G{br},$J{br}+($C{br}-$F{br})/($G{br}-$F{br})*($K{br}-$J{br}),"
                f"IF($C{br}<=$H{br},$K{br}+($C{br}-$G{br})/($H{br}-$G{br})*($L{br}-$K{br}),"
                f"$L{br}+($C{br}-$H{br})/($I{br}-$H{br})*($M{br}-$L{br}))))))"
            ),
        )
        br += 1
    last_metric = br - 1
    ws_b.cell(row=br, column=13, value="Group means (group-weighted B, ADR-0006):")
    gm_cells = []
    for gi, group in enumerate(("scale", "unit_economics", "momentum")):
        gr = br + 1 + gi
        ws_b.cell(row=gr, column=13, value=group)
        ws_b.cell(
            row=gr,
            column=14,
            value=f'=AVERAGEIF($B$2:$B${last_metric},"{group}",$N$2:$N${last_metric})',
        )
        gm_cells.append(f"$N${gr}")
    b_row = br + 4
    ws_b.cell(row=b_row, column=13, value="B =")
    ws_b.cell(row=b_row, column=14, value=f"=AVERAGE({','.join(gm_cells)})").font = Font(bold=True)
    b_cell = f"Business!$N${b_row}"
    _autofit(ws_b, {1: 24, 2: 15})

    # --- Powers (Benefit/Barrier; strength = the weaker; all 7 in scope; P = mean over all 7) ---
    ws_p = wb.create_sheet("Powers")
    _header(ws_p, 1, ["Power", "Benefit", "Barrier", "Strength (weaker)", "Value"])

    def _rank(col: str, row: int) -> str:
        c = f"${col}{row}"
        return f'IF({c}="None",0,IF({c}="Emerging",1,IF({c}="Established",2,3)))'

    pr = 2
    for power in registry.powers:
        benefit, barrier = POWER_ASSESSMENT[power.key]  # never N/A (Methodology v1.1 §8)
        ws_p.cell(row=pr, column=1, value=power.key)
        ws_p.cell(row=pr, column=2, value=benefit)
        ws_p.cell(row=pr, column=3, value=barrier)
        # Strength = the weaker of Benefit/Barrier (Helmer both-required), live.
        ws_p.cell(
            row=pr,
            column=4,
            value=f"=IF(({_rank('B', pr)})<=({_rank('C', pr)}),$B{pr},$C{pr})",
        )
        ws_p.cell(
            row=pr,
            column=5,
            value=(
                f'=IF($D{pr}="","",IF($D{pr}="None",0,IF($D{pr}="Emerging",0.4,'
                f'IF($D{pr}="Established",0.7,IF($D{pr}="Wide",1,"")))))'
            ),
        )
        pr += 1
    last_power = pr - 1
    ws_p.cell(row=pr, column=4, value="P =")
    ws_p.cell(row=pr, column=5, value=f"=AVERAGE($E$2:$E${last_power})").font = Font(bold=True)
    p_cell = f"Powers!$E${pr}"
    _autofit(ws_p, {1: 22, 4: 18})

    # --- Composite (L, V live; triad ordinal) ---
    ws_v = wb.create_sheet("Composite")
    ws_v.cell(
        row=1, column=1, value="ATLAS Composite — V = θ_B·B + θ_P·P + θ_L·L"
    ).font = _TITLE_FONT
    qm_all = ",".join(module_qm_cell[m.key] for m in registry.modules)
    qm_crit = ",".join(module_qm_cell[k] for k in CRITICAL_MODULES_FOR_L)
    triad = result["triad"]
    rows = [
        ("alpha_L", ALPHA_L, None),
        ("L weighted term (mean q_m)", None, f"=AVERAGE({qm_all})"),
        ("L min term (critical modules)", None, f"=MIN({qm_crit})"),
        ("L", None, "=$B3*$C4+(1-$B3)*$C5"),
        ("B", None, f"={b_cell}"),
        ("P", None, f"={p_cell}"),
        ("theta_B", THETA["B"], None),
        ("theta_P", THETA["P"], None),
        ("theta_L", THETA["L"], None),
        ("V", None, "=$B9*$C7+$B10*$C8+$B11*$C6"),
        ("V (display 0–100)", None, "=$C12*100"),
        ("", None, None),
        ("— Platform Power Triad (ordinal, ADR-0007) —", None, None),
        ("Economic Value", None, triad["economic_value"]["rating"]),
        ("Perceived Value", None, triad["perceived_value"]["rating"]),
        ("Defence Value", None, triad["defence_value"]["rating"]),
    ]
    for col, text in ((1, "Item"), (2, "Coefficient"), (3, "Value")):
        cell = ws_v.cell(row=2, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for i, (label, coeff, formula) in enumerate(rows, start=3):
        ws_v.cell(row=i, column=1, value=label)
        if coeff is not None:
            ws_v.cell(row=i, column=2, value=coeff)
        if formula is not None:
            ws_v.cell(row=i, column=3, value=formula)
    ws_v.cell(row=12, column=1).font = Font(bold=True)
    ws_v.cell(row=12, column=3).font = Font(bold=True)
    _autofit(ws_v, {1: 34, 2: 14, 3: 14})

    # --- Overview (read me) ---
    ws_o = wb.create_sheet("Overview", 0)
    ws_o.cell(row=1, column=1, value="ATLAS Golden Master — Meridian Securities").font = _TITLE_FONT
    c = result["composite"]
    t = result["triad"]
    lines = [
        f"Subject: {SUBJECT}",
        f"Methodology: v{METHODOLOGY_VERSION}  |  Status: {FIXTURE_STATUS}",
        "",
        "This workbook computes one complete assessment with LIVE formulas. Edit a Level on the",
        "Subcomponents sheet (or set a subcomponent/metric State: NOT_APPLICABLE / NOT_ASSESSED)",
        "and q_m, L, B, P, V recompute. B is a group-weighted mean (scale/unit-econ/momentum,",
        "ADR-0006); powers carry Benefit + Barrier and the weaker side gates (Helmer). All 7",
        "powers are always in scope — N/A is not a legal power state (Methodology v1.1 §8). Gate",
        "bands and the triad are rule-based/derived (computed values, regenerated by Python).",
        "",
        "DRAFT judgement calls for John (GRS-0003 review): subcomponent levels/evidence; critical",
        "flags; the Network Economies / Counter-Positioning strengths (now real LOW levels, not",
        "N/A); the draft coefficients (α, δ, λ, θ, group weights); the strength encoding",
        "(ADR-0004); the gate rule (ADR-0003); B/P range comparability (ADR-0005). Edit, save,",
        "then run",
        "scripts/regen_golden_master_json.py to refresh the JSON fixture.",
        "",
        f"Headline: B={c['B']}  P={c['P']}  L={c['L']}  →  V={c['V']} "
        f"(display {result['two_track']['continuous']['V_display_0_100']}).",
        f"Triad: Economic {t['economic_value']['rating']}  ·  Perceived "
        f"{t['perceived_value']['rating']}  ·  Defence {t['defence_value']['rating']}.",
    ]
    for i, text in enumerate(lines, start=3):
        cell = ws_o.cell(row=i, column=1, value=text)
        if text.startswith(("DRAFT", "This workbook")):
            cell.font = _NOTE_FONT
    ws_o.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _autofit(ws: Worksheet, widths: dict[int, int]) -> None:
    from openpyxl.utils import get_column_letter

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def main() -> None:
    registry = load_registry()
    result = compute(registry)
    write_json(result)
    print(f"wrote {_JSON_OUT.relative_to(_ROOT)}")
    write_workbook(registry, result, _XLSX_OUT)
    print(f"wrote {_XLSX_OUT.relative_to(_ROOT)}")
    c = result["composite"]
    print(
        f"\nMeridian Securities: B={c['B']} P={c['P']} L={c['L']} -> V={c['V']} "
        f"(display {result['two_track']['continuous']['V_display_0_100']})"
    )


if __name__ == "__main__":
    main()
