"""Shared plumbing for the GTM registry import scripts (GRS-0193).

Operator-run, not part of any request path. Each importer reads a file whose path comes from an
argument or an environment variable (never a path committed into the script), upserts through the
repository, and prints its summary as JSON so a re-run is auditable.
"""

from __future__ import annotations

import json
import os
from collections.abc import Callable, Iterator, Sequence
from contextlib import contextmanager
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from grassmarket.config import get_settings
from grassmarket.data.database import make_engine, make_session_factory
from grassmarket.data.repository import Repository
from grassmarket.gtm import ImportSummary


def resolve_path(argv: Sequence[str], *, env_var: str, what: str) -> Path:
    """The source file: `argv[1]` if given, else `$env_var`. Fails loud with a usage message rather
    than falling back to a guessed location."""
    candidate = argv[1] if len(argv) > 1 else os.environ.get(env_var)
    if not candidate:
        raise SystemExit(f"Usage: {argv[0]} <path-to-{what}>  (or set {env_var})")
    path = Path(candidate).expanduser()
    if not path.is_file():
        raise SystemExit(f"{what} not found at {path}.")
    return path


def import_date() -> date:
    """The run date, overridable with GTM_IMPORT_DATE so a re-import can be stamped consistently."""
    override = os.environ.get("GTM_IMPORT_DATE")
    if override:
        return date.fromisoformat(override)
    return datetime.now(UTC).date()


@contextmanager
def repository() -> Iterator[Repository]:
    """A repository over one committed transaction: the whole import lands, or none of it does."""
    factory = make_session_factory(make_engine(get_settings().database_url))
    session = factory()
    try:
        yield Repository(session)
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def run(build: Callable[[Repository], ImportSummary]) -> None:
    """Run an importer and print its summary. Any RowError propagates and aborts the transaction —
    a partial import that silently dropped rows is worse than no import (#3)."""
    with repository() as repo:
        summary = build(repo)
    print(json.dumps(summary.as_dict(), indent=2))


def read_xlsx_rows(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    """Header-keyed rows from a worksheet. Kept here rather than in `grassmarket.gtm` so the parsing
    layer stays free of file IO and its tests need no spreadsheet."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet is not None else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise SystemExit(f"Sheet {worksheet.title!r} in {path} is empty.")
        keys = [str(h).strip() if h is not None else "" for h in header]
        return [dict(zip(keys, row, strict=False)) for row in rows]
    finally:
        workbook.close()
