"""Import the Barclays Live influencer workbook into the GTM registry (GRS-0193).

    uv run python scripts/import_barclays_influencer.py \
        data/gtm/sources/barclays-influencer-map.xlsx

The workbook is the worked example the method came from: an Influencer Map tab of ranked analysts
(LSEG-derived, so unverified) and a Target Owners tab of ownership rows carrying a human's
web-research verification. Both land against one Barclays target, keeping the two-source split
GRS-0194 depends on: an owner row imports as verified only when its Verification column says so.

The tabs carry two header rows of title and provenance above the real header, so each is read by
locating its header row rather than assuming a fixed offset.
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

from _gtm_import import import_date, repository, resolve_path
from bcap_contracts.entities import RegistryTarget

from grassmarket.data.repository import Repository
from grassmarket.gtm import (
    ImportSummary,
    parse_barclays_analyst_row,
    parse_barclays_owner_row,
)

TARGET_ID = "lseg-barclays"
ANALYST_SHEET = "Influencer Map"
OWNER_SHEET = "Target - Barclays Live Owners"
ANALYST_HEADER = "Rank"
OWNER_HEADER = "Name"


def _rows_below_header(path: Path, sheet: str, header_cell: str) -> list[dict[str, Any]]:
    """Rows keyed by the header found at the first row whose first cell is `header_cell`."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    for index, row in enumerate(rows):
        if row and str(row[0]).strip() == header_cell:
            keys = [str(h).strip() if h is not None else "" for h in row]
            return [
                dict(zip(keys, r, strict=False))
                for r in rows[index + 1 :]
                if r and r[0] is not None
            ]
    raise SystemExit(f"No header row starting {header_cell!r} found in sheet {sheet!r}.")


def _barclays_target(on: date) -> RegistryTarget:
    return RegistryTarget(
        target_id=TARGET_ID,
        name="Barclays",
        aliases=("Barclays Investment Bank", "Barclays Live", "Barclays Research"),
        domain="barclays.com",
        segment="Sell-side research",
        country="United Kingdom",
        ric=None,
        # Contributor 10333, established by the GRS-0200 pull; GRS-0194 keys the map generator off
        # this, which is why it is set here rather than left to inference.
        ctb_id=10333,
        source="barclays-influencer-map",
        imported_on=on,
    )


def main() -> None:
    path = resolve_path(sys.argv, env_var="GTM_BARCLAYS_XLSX", what="barclays-influencer-map.xlsx")
    on = import_date()
    summary = ImportSummary(source="barclays-influencer-map")
    with repository() as repo:
        _upsert(repo, path, on, summary)
    print(json.dumps(summary.as_dict(), indent=2))


def _upsert(repo: Repository, path: Path, on: date, summary: ImportSummary) -> None:
    repo.upsert_registry_target(_barclays_target(on))
    summary.targets_upserted += 1
    for row in _rows_below_header(path, ANALYST_SHEET, ANALYST_HEADER):
        summary.rows_read += 1
        repo.upsert_registry_contact(
            parse_barclays_analyst_row(row, target_id=TARGET_ID, imported_on=on)
        )
        summary.contacts_upserted += 1
    for row in _rows_below_header(path, OWNER_SHEET, OWNER_HEADER):
        summary.rows_read += 1
        contact = parse_barclays_owner_row(row, target_id=TARGET_ID, imported_on=on)
        repo.upsert_registry_contact(contact)
        summary.contacts_upserted += 1
        if not contact.verified:
            summary.skip(f"{contact.full_name}: ownership row not cleanly verified, flagged")


if __name__ == "__main__":
    main()
