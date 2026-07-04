"""Regenerate tests/fixtures/golden_master.json from the edited workbook (GRS-0003).

Workflow: John opens fixtures/golden-master.xlsx in Excel, adjusts subcomponent Levels / Evidence
/ States (and, if needed, metric raws or power strengths), saves, then runs this script. It reads
those INPUTS back from the workbook and re-runs the same reference computation
(`build_golden_master.compute`) so the JSON stays the single oracle — one compute path, no drift.

Coefficients remain the draft constants in build_golden_master (they come from the elicitation
panel, not manual workbook edits). Run from the repo root: `uv run python scripts/regen_...`.
"""

from __future__ import annotations

from pathlib import Path

import build_golden_master as bgm
import openpyxl

_ROOT = Path(__file__).resolve().parents[1]
_XLSX = _ROOT / "fixtures" / "golden-master.xlsx"


def _read_inputs() -> (
    tuple[
        dict[str, tuple[str, str | None]], dict[str, float | str], dict[str, tuple[str, str] | str]
    ]
):
    wb = openpyxl.load_workbook(_XLSX, data_only=False)

    subs: dict[str, tuple[str, str | None]] = {}
    ws = wb["Subcomponents"]
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, 2).value
        if not key:
            continue
        level = (ws.cell(row, 4).value or "").strip()
        evidence = (ws.cell(row, 6).value or "").strip() or None
        state = (ws.cell(row, 7).value or "").strip()
        subs[key] = (state, None) if state else (level, evidence)

    metrics: dict[str, float | str] = {}
    wb_b = wb["Business"]
    for row in range(2, wb_b.max_row + 1):
        key = wb_b.cell(row, 1).value
        raw = wb_b.cell(row, 3).value  # Raw is column C (Group is column B)
        if not key:
            continue
        metrics[key] = float(raw) if isinstance(raw, int | float) else str(raw).strip()

    powers: dict[str, tuple[str, str] | str] = {}
    wb_p = wb["Powers"]
    for row in range(2, wb_p.max_row + 1):
        key = wb_p.cell(row, 1).value
        if not key:
            continue
        state = (wb_p.cell(row, 6).value or "").strip()
        if state:
            powers[key] = state
        else:
            benefit = (wb_p.cell(row, 2).value or "").strip()
            barrier = (wb_p.cell(row, 3).value or "").strip()
            powers[key] = (benefit, barrier)
    return subs, metrics, powers


def main() -> None:
    subs, metrics, powers = _read_inputs()
    # Inject the workbook-read inputs into the reference generator, then recompute + rewrite JSON.
    bgm.SUBCOMPONENTS = subs
    bgm.METRIC_RAW = metrics
    bgm.POWER_ASSESSMENT = powers
    result = bgm.compute(bgm.load_registry())
    bgm.write_json(result)
    c = result["composite"]
    print(f"regenerated golden_master.json from workbook: V={c['V']} (display {c['V'] * 100:.4f})")


if __name__ == "__main__":
    main()
